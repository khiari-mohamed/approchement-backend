"""
Export service for generating Excel and PDF reports
Production-ready with proper formatting and Tunisian standards
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from datetime import datetime
import os

class ExportService:
    """Production-ready export service for reconciliation reports"""
    
    def __init__(self, storage_path: str = "storage/reports"):
        self.storage_path = storage_path
        os.makedirs(storage_path, exist_ok=True)
    
    def export_to_excel(self, reconciliation_data: dict, filename: str = None) -> str:
        """Export reconciliation to Excel with multiple sheets"""
        if not filename:
            filename = f"rapprochement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(self.storage_path, filename)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary
        self._create_summary_sheet(wb, reconciliation_data)
        
        # Sheet 2: Matches
        self._create_matches_sheet(wb, reconciliation_data)
        
        # Sheet 3: Suspense Items
        self._create_suspense_sheet(wb, reconciliation_data)
        
        # Sheet 4: Regularization Entries
        if "regularization_entries" in reconciliation_data:
            self._create_regularization_sheet(wb, reconciliation_data)
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, data: dict):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Résumé")
        summary = data.get("summary", {})
        
        # Title
        ws['A1'] = "ÉTAT DE RAPPROCHEMENT BANCAIRE"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A2'] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(size=10)
        
        # Company info (if available)
        ws['A4'] = "Entreprise:"
        ws['B4'] = data.get("company_name", "N/A")
        ws['A5'] = "Période:"
        ws['B5'] = data.get("period", "N/A")
        
        # Summary metrics
        row = 7
        metrics = [
            ("Total Bancaire", summary.get("bank_total", 0), "TND"),
            ("Total Comptable", summary.get("accounting_total", 0), "TND"),
            ("Écart Initial", summary.get("initial_gap", 0), "TND"),
            ("", "", ""),
            ("Transactions Rapprochées", summary.get("matched_count", 0), ""),
            ("Transactions en Suspens", summary.get("suspense_count", 0), ""),
            ("Taux de Couverture", f"{summary.get('coverage_ratio', 0) * 100:.1f}", "%"),
            ("", "", ""),
            ("Écart Résiduel", summary.get("residual_gap", 0), "TND"),
        ]
        
        for label, value, unit in metrics:
            if label:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                if isinstance(value, (int, float)):
                    ws[f'B{row}'].number_format = '#,##0.000'
                ws[f'C{row}'] = unit
            row += 1
        
        # Styling
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_matches_sheet(self, wb: Workbook, data: dict):
        """Create matches sheet with all reconciled transactions"""
        ws = wb.create_sheet("Rapprochements")
        
        # Headers
        headers = ["N° R", "Date Banque", "Libellé Banque", "Date Compta", 
                  "Libellé Compta", "Montant", "Règle", "Score", "Statut"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        matches = data.get("matches", [])
        for match in matches:
            bank_tx = match.get("bankTx", {})
            acc_tx = match.get("accountingTx", {})
            
            ws.append([
                match.get("reconId", ""),
                bank_tx.get("date", ""),
                bank_tx.get("description", ""),
                acc_tx.get("date", "") if acc_tx else "",
                acc_tx.get("description", "") if acc_tx else "",
                bank_tx.get("amount", 0),
                match.get("rule", ""),
                f"{match.get('score', 0) * 100:.0f}%",
                match.get("status", "")
            ])
        
        # Format amount column
        for row in range(2, len(matches) + 2):
            ws.cell(row=row, column=6).number_format = '#,##0.000'
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_suspense_sheet(self, wb: Workbook, data: dict):
        """Create suspense items sheet"""
        ws = wb.create_sheet("Suspens")
        
        # Headers
        headers = ["Type", "Date", "Libellé", "Montant", "Catégorie Suggérée", 
                  "Compte PCN", "Raison"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        suspense = data.get("suspense", [])
        for item in suspense:
            tx = item.get("transaction", {})
            ws.append([
                "Bancaire" if item.get("type") == "bank" else "Comptable",
                tx.get("date", ""),
                tx.get("description", ""),
                tx.get("amount", 0),
                item.get("suggestedCategory", ""),
                item.get("suggestedAccount", ""),
                item.get("reason", "")
            ])
        
        # Format amount column
        for row in range(2, len(suspense) + 2):
            ws.cell(row=row, column=4).number_format = '#,##0.000'
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_regularization_sheet(self, wb: Workbook, data: dict):
        """Create regularization entries sheet"""
        ws = wb.create_sheet("Écritures de Régularisation")
        
        # Headers
        headers = ["N° Écriture", "Date", "Compte", "Libellé Compte", 
                  "Description", "Débit", "Crédit"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        entries = data.get("regularization_entries", [])
        for entry in entries:
            for line in entry.get("lines", []):
                ws.append([
                    entry.get("entry_number", ""),
                    entry.get("date", ""),
                    line.get("account_code", ""),
                    line.get("account_name", ""),
                    line.get("description", ""),
                    line.get("debit", 0),
                    line.get("credit", 0)
                ])
        
        # Format amount columns
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=6).number_format = '#,##0.000'
            ws.cell(row=row, column=7).number_format = '#,##0.000'
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def export_to_pdf(self, reconciliation_data: dict, filename: str = None) -> str:
        """Export reconciliation to PDF report"""
        if not filename:
            filename = f"rapprochement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = os.path.join(self.storage_path, filename)
        
        # Create PDF
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("ÉTAT DE RAPPROCHEMENT BANCAIRE", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Summary section
        summary = reconciliation_data.get("summary", {})
        summary_data = [
            ["Métrique", "Valeur"],
            ["Total Bancaire", f"{summary.get('bank_total', 0):,.3f} TND"],
            ["Total Comptable", f"{summary.get('accounting_total', 0):,.3f} TND"],
            ["Écart Initial", f"{summary.get('initial_gap', 0):,.3f} TND"],
            ["Transactions Rapprochées", str(summary.get('matched_count', 0))],
            ["Transactions en Suspens", str(summary.get('suspense_count', 0))],
            ["Taux de Couverture", f"{summary.get('coverage_ratio', 0) * 100:.1f}%"],
            ["Écart Résiduel", f"{summary.get('residual_gap', 0):,.3f} TND"],
        ]
        
        summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(PageBreak())
        
        # Matches section
        elements.append(Paragraph("Détail des Rapprochements", styles['Heading2']))
        elements.append(Spacer(1, 0.3*cm))
        
        matches = reconciliation_data.get("matches", [])[:50]  # Limit for PDF
        if matches:
            match_data = [["N° R", "Date", "Libellé", "Montant", "Règle", "Score"]]
            for match in matches:
                bank_tx = match.get("bankTx", {})
                match_data.append([
                    match.get("reconId", "")[:10],
                    bank_tx.get("date", ""),
                    bank_tx.get("description", "")[:40],
                    f"{bank_tx.get('amount', 0):,.2f}",
                    match.get("rule", "")[:15],
                    f"{match.get('score', 0) * 100:.0f}%"
                ])
            
            match_table = Table(match_data, colWidths=[3*cm, 3*cm, 8*cm, 3*cm, 3*cm, 2*cm])
            match_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            
            elements.append(match_table)
        else:
            elements.append(Paragraph("Aucun rapprochement trouvé", styles['Normal']))
        
        elements.append(PageBreak())
        
        # Suspense section
        elements.append(Paragraph("Opérations en Suspens", styles['Heading2']))
        elements.append(Spacer(1, 0.3*cm))
        
        suspense = reconciliation_data.get("suspense", [])[:100]  # Limit for PDF
        if suspense:
            suspense_data = [["Type", "Date", "Libellé", "Montant", "Raison"]]
            for item in suspense:
                tx = item.get("transaction", {})
                suspense_data.append([
                    "Bancaire" if item.get("type") == "bank" else "Comptable",
                    tx.get("date", ""),
                    tx.get("description", "")[:45],
                    f"{tx.get('amount', 0):,.2f}",
                    item.get("reason", "")
                ])
            
            suspense_table = Table(suspense_data, colWidths=[2.5*cm, 2.5*cm, 8*cm, 3*cm, 6*cm])
            suspense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C65911')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            
            elements.append(suspense_table)
        else:
            elements.append(Paragraph("Aucune opération en suspens", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        return filepath
    
    def export_regularization_to_csv(self, entries: list, filename: str = None) -> str:
        """Export regularization entries to CSV for accounting software import"""
        if not filename:
            filename = f"ecritures_reg_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = os.path.join(self.storage_path, filename)
        
        # Flatten entries to rows
        rows = []
        for entry in entries:
            for line in entry.get("lines", []):
                rows.append({
                    "Journal": "OD",
                    "N° Écriture": entry.get("entry_number"),
                    "Date": entry.get("date"),
                    "Compte": line.get("account_code"),
                    "Libellé Compte": line.get("account_name"),
                    "Description": line.get("description"),
                    "Débit": line.get("debit", 0),
                    "Crédit": line.get("credit", 0),
                    "Devise": "TND",
                    "N° Pièce": entry.get("entry_number")
                })
        
        df = pd.DataFrame(rows)
        df.to_csv(filepath, index=False, encoding='utf-8-sig', sep=';')
        return filepath
